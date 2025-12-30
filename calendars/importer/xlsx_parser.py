"""XLSX schedule parser."""
from __future__ import annotations

from typing import Any, Dict, List

from .base import ScheduleParser
from .model import ScheduleItem


class XLSXParser(ScheduleParser):
    """Parser for Excel (.xlsx) schedule files."""

    def parse(self, path: str) -> List[ScheduleItem]:
        """Parse schedule items from XLSX file.

        Args:
            path: Path to XLSX file

        Returns:
            List of ScheduleItem objects

        Raises:
            RuntimeError: If openpyxl is not installed
        """
        try:
            import openpyxl  # type: ignore
        except Exception as e:  # pragma: no cover - optional
            raise RuntimeError("openpyxl is required to parse .xlsx files. Try: python3 -m pip install openpyxl") from e

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # Expect a header row; map columns by name
        headers: Dict[int, str] = {}
        items: List[ScheduleItem] = []

        for i, cell in enumerate(ws[1], start=1):
            headers[i] = str(cell.value or '').strip()

        for r in ws.iter_rows(min_row=2):
            # Convert cell values to strings for uniform handling
            row: Dict[str, Any] = {}
            for i, cell in enumerate(r, start=1):
                row[headers.get(i, f'col{i}')] = str(cell.value) if cell.value is not None else ''
            item = self._row_to_schedule_item(row)
            if item:
                items.append(item)

        return items


# Backward compatibility function
def parse_xlsx(path: str) -> List[ScheduleItem]:
    """Parse schedule items from XLSX file.

    Args:
        path: Path to XLSX file

    Returns:
        List of ScheduleItem objects
    """
    return XLSXParser().parse(path)
