"""Tests for SheetGenerator internals."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from sheets.constants import MAX_COLUMN_WIDTH, MIN_COLUMN_WIDTH
from sheets.generator import (
    SheetGenerator,
    generate_from_yaml,
    generate_xlsx,
    validate_workbook,
)
from sheets.schema import HeaderStyle, SheetMetadata, SheetTab, SheetWorkbook


def _make_workbook(sheets=None) -> SheetWorkbook:
    if sheets is None:
        sheets = [SheetTab(name="Sheet1", headers=["A", "B"], rows=[[1, 2], [3, 4]])]
    return SheetWorkbook(metadata=SheetMetadata(title="Test"), sheets=sheets)


class TestNormalizeColor(unittest.TestCase):
    def setUp(self):
        self.gen = SheetGenerator()

    def test_with_hash_prefix(self):
        self.assertEqual(self.gen._normalize_color("#2D3A4F"), "FF2D3A4F")

    def test_without_hash(self):
        self.assertEqual(self.gen._normalize_color("2D3A4F"), "FF2D3A4F")

    def test_already_argb(self):
        self.assertEqual(self.gen._normalize_color("FF2D3A4F"), "FF2D3A4F")

    def test_lowercase_upcased(self):
        self.assertEqual(self.gen._normalize_color("2d3a4f"), "FF2D3A4F")


class TestColumnWidthClamping(unittest.TestCase):
    def setUp(self):
        self.gen = SheetGenerator()

    def test_clamps_to_min(self):
        # A single short value should hit the MIN floor
        width = self.gen._calculate_column_width(["X"])
        self.assertGreaterEqual(width, MIN_COLUMN_WIDTH)

    def test_clamps_to_max(self):
        # A very long value should be capped at MAX
        long_val = "A" * 200
        width = self.gen._calculate_column_width([long_val])
        self.assertEqual(width, MAX_COLUMN_WIDTH)

    def test_auto_fit_from_content(self):
        values = ["Col", "short", "a longer value here"]
        width = self.gen._calculate_column_width(values)
        self.assertGreaterEqual(width, MIN_COLUMN_WIDTH)
        self.assertLessEqual(width, MAX_COLUMN_WIDTH)

    def test_explicit_column_widths_used_when_provided(self):
        sheet = SheetTab(
            name="S",
            headers=["A", "B"],
            rows=[],
            column_widths=[15, 25],
        )
        widths = self.gen._resolve_column_widths(sheet)
        self.assertEqual(widths, [15, 25])

    def test_auto_widths_when_not_provided(self):
        sheet = SheetTab(name="S", headers=["A", "B"], rows=[["x", "y"]])
        widths = self.gen._resolve_column_widths(sheet)
        self.assertEqual(len(widths), 2)
        for w in widths:
            self.assertGreaterEqual(w, MIN_COLUMN_WIDTH)
            self.assertLessEqual(w, MAX_COLUMN_WIDTH)


class TestFreezePanes(unittest.TestCase):
    def setUp(self):
        self.gen = SheetGenerator()

    def test_no_freeze_when_both_zero(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        self.gen._apply_freeze_panes(ws, 0, 0)
        self.assertIsNone(ws.freeze_panes)

    def test_freeze_one_row(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        self.gen._apply_freeze_panes(ws, 1, 0)
        self.assertEqual(ws.freeze_panes, "A2")

    def test_freeze_row_and_col(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        self.gen._apply_freeze_panes(ws, 2, 1)
        self.assertEqual(ws.freeze_panes, "B3")


class TestGenerateXlsx(unittest.TestCase):
    def test_generate_creates_file_with_expected_sheets(self):
        wb_def = _make_workbook()
        with tempfile.TemporaryDirectory() as tmpdir:
            out = str(Path(tmpdir) / "out.xlsx")
            result = generate_xlsx(wb_def, out)
            self.assertEqual(result, out)
            self.assertTrue(Path(out).exists())
            wb = load_workbook(out)
            self.assertIn("Sheet1", wb.sheetnames)

    def test_generate_from_yaml_end_to_end(self):
        import yaml
        wb_def = {
            "title": "E2E",
            "sheets": [
                {"name": "Data", "headers": ["X", "Y"], "rows": [[1, 2]]}
            ],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            yaml_path = str(Path(tmpdir) / "wb.yaml")
            with open(yaml_path, "w") as f:
                yaml.dump(wb_def, f)
            out = str(Path(tmpdir) / "out.xlsx")
            result = generate_from_yaml(yaml_path, out)
            self.assertTrue(Path(result).exists())
            wb = load_workbook(result)
            self.assertEqual(wb.sheetnames, ["Data"])
            ws = wb["Data"]
            self.assertEqual(ws.cell(1, 1).value, "X")
            self.assertEqual(ws.cell(2, 1).value, 1)

    def test_multi_tab_workbook_all_sheets_present(self):
        wb_def = SheetWorkbook(
            metadata=SheetMetadata(title="Multi"),
            sheets=[
                SheetTab(name="Alpha", headers=["A"], rows=[]),
                SheetTab(name="Beta", headers=["B"], rows=[]),
            ],
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            out = str(Path(tmpdir) / "multi.xlsx")
            generate_xlsx(wb_def, out)
            wb = load_workbook(out)
            self.assertEqual(set(wb.sheetnames), {"Alpha", "Beta"})

    def test_header_styling_applied(self):
        style = HeaderStyle(bg_color="#FF0000", text_color="#FFFFFF", bold=True, font_size=14)
        wb_def = SheetWorkbook(
            metadata=SheetMetadata(title="Styled"),
            sheets=[SheetTab(name="S", headers=["Col"], rows=[], header_style=style)],
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            out = str(Path(tmpdir) / "styled.xlsx")
            generate_xlsx(wb_def, out)
            wb = load_workbook(out)
            ws = wb["S"]
            cell = ws.cell(1, 1)
            # Header should be bold
            self.assertTrue(cell.font.bold)
            self.assertEqual(cell.font.size, 14)

    def test_alternating_rows_applied(self):
        wb_def = SheetWorkbook(
            metadata=SheetMetadata(title="Zebra"),
            sheets=[SheetTab(name="Z", headers=["A"], rows=[["x"], ["y"], ["z"]], alternating_rows=True)],
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            out = str(Path(tmpdir) / "zebra.xlsx")
            generate_xlsx(wb_def, out)
            wb = load_workbook(out)
            ws = wb["Z"]
            # Row 2 (first data row, index 0 = even) gets ALTERNATING_ROW_COLOR_1
            # Row 3 (second data row, index 1 = odd) gets ALTERNATING_ROW_COLOR_2
            fill_row2 = ws.cell(2, 1).fill.fgColor.rgb
            fill_row3 = ws.cell(3, 1).fill.fgColor.rgb
            self.assertNotEqual(fill_row2, fill_row3)


class TestValidateWorkbook(unittest.TestCase):
    """validate_workbook catches definitions openpyxl cannot write.

    Regression guard: `sheets validate` previously reported OK for a workbook
    whose sheet name Excel forbids, and `sheets generate` then failed at write
    time — a validate pass that did not mean the file could be produced.
    """

    def test_clean_workbook_has_no_problems(self) -> None:
        self.assertEqual(validate_workbook(_make_workbook()), [])

    def test_forbidden_character_is_reported(self) -> None:
        wb = _make_workbook([SheetTab(name="Q1/Q2", headers=["A"], rows=[[1]])])
        problems = validate_workbook(wb)
        self.assertEqual(len(problems), 1)
        self.assertIn("/", problems[0])

    def test_every_forbidden_character_is_caught(self) -> None:
        for char in ["\\", "/", "*", "?", ":", "[", "]"]:
            with self.subTest(char=char):
                wb = _make_workbook(
                    [SheetTab(name=f"A{char}B", headers=["A"], rows=[[1]])]
                )
                self.assertTrue(validate_workbook(wb), f"{char!r} not caught")

    def test_forbidden_character_actually_breaks_openpyxl(self) -> None:
        # Anchors the constant list to real openpyxl behaviour rather than a
        # guess: if openpyxl ever stops rejecting one of these, this fails.
        wb = _make_workbook([SheetTab(name="Q1/Q2", headers=["A"], rows=[[1]])])
        with tempfile.TemporaryDirectory() as tmp:
            out = str(Path(tmp) / "out.xlsx")
            with self.assertRaises(ValueError):
                generate_xlsx(wb, out)

    def test_overlong_name_is_reported(self) -> None:
        wb = _make_workbook([SheetTab(name="A" * 32, headers=["A"], rows=[[1]])])
        problems = validate_workbook(wb)
        self.assertEqual(len(problems), 1)
        self.assertIn("32", problems[0])

    def test_name_at_limit_is_allowed(self) -> None:
        wb = _make_workbook([SheetTab(name="A" * 31, headers=["A"], rows=[[1]])])
        self.assertEqual(validate_workbook(wb), [])

    def test_empty_name_is_reported(self) -> None:
        wb = _make_workbook([SheetTab(name="", headers=["A"], rows=[[1]])])
        self.assertTrue(validate_workbook(wb))

    def test_duplicate_names_are_reported(self) -> None:
        wb = _make_workbook([
            SheetTab(name="Data", headers=["A"], rows=[[1]]),
            SheetTab(name="Data", headers=["A"], rows=[[2]]),
        ])
        problems = validate_workbook(wb)
        self.assertTrue(any("duplicate" in p for p in problems))

    def test_duplicate_names_differing_only_by_case_are_reported(self) -> None:
        # Excel compares sheet names case-insensitively.
        wb = _make_workbook([
            SheetTab(name="Data", headers=["A"], rows=[[1]]),
            SheetTab(name="DATA", headers=["A"], rows=[[2]]),
        ])
        problems = validate_workbook(wb)
        self.assertTrue(any("duplicate" in p for p in problems))

    def test_multiple_problems_all_reported(self) -> None:
        wb = _make_workbook([
            SheetTab(name="A/B", headers=["A"], rows=[[1]]),
            SheetTab(name="C" * 40, headers=["A"], rows=[[1]]),
        ])
        self.assertEqual(len(validate_workbook(wb)), 2)


class TestUncoveredGeneratorPaths(unittest.TestCase):
    """Branches the initial test pass missed."""

    def test_alternating_rows_disabled_leaves_cells_unfilled(self) -> None:
        wb = SheetWorkbook(
            metadata=SheetMetadata(title="T"),
            sheets=[SheetTab(
                name="Plain", headers=["A"], rows=[[1], [2]],
                alternating_rows=False,
            )],
        )
        with tempfile.TemporaryDirectory() as tmp:
            out = str(Path(tmp) / "plain.xlsx")
            generate_xlsx(wb, out)
            ws = load_workbook(out)["Plain"]
            self.assertEqual(ws.cell(2, 1).fill.fill_type, None)
            self.assertEqual(ws.cell(3, 1).fill.fill_type, None)

    def test_column_width_skips_none_values(self) -> None:
        gen = SheetGenerator()
        width = gen._calculate_column_width([None, "hello"])
        self.assertGreaterEqual(width, MIN_COLUMN_WIDTH)
        self.assertLessEqual(width, MAX_COLUMN_WIDTH)

    def test_row_shorter_than_headers_is_written(self) -> None:
        wb = SheetWorkbook(
            metadata=SheetMetadata(title="T"),
            sheets=[SheetTab(name="Short", headers=["A", "B", "C"], rows=[["x"]])],
        )
        with tempfile.TemporaryDirectory() as tmp:
            out = str(Path(tmp) / "short.xlsx")
            generate_xlsx(wb, out)
            ws = load_workbook(out)["Short"]
            self.assertEqual(ws.cell(2, 1).value, "x")
            self.assertIsNone(ws.cell(2, 3).value)

    def test_author_is_written_to_workbook_properties(self) -> None:
        wb = SheetWorkbook(
            metadata=SheetMetadata(title="T", author="Alice"),
            sheets=[SheetTab(name="S", headers=["A"], rows=[[1]])],
        )
        with tempfile.TemporaryDirectory() as tmp:
            out = str(Path(tmp) / "authored.xlsx")
            generate_xlsx(wb, out)
            self.assertEqual(load_workbook(out).properties.creator, "Alice")


class TestPRFeedbackFixes(unittest.TestCase):
    """Regressions found in review of PR #238."""

    def test_short_content_reaches_the_documented_minimum(self) -> None:
        # Previously seeded from MIN_COLUMN_WIDTH then added +2, so the real
        # floor was MIN + 2 and the documented lower bound was unreachable.
        gen = SheetGenerator()
        self.assertEqual(gen._calculate_column_width(["X"]), MIN_COLUMN_WIDTH)
        self.assertEqual(gen._calculate_column_width([""]), MIN_COLUMN_WIDTH)
        self.assertEqual(gen._calculate_column_width([None]), MIN_COLUMN_WIDTH)

    def test_content_longer_than_minimum_still_gets_buffer(self) -> None:
        gen = SheetGenerator()
        value = "A" * (MIN_COLUMN_WIDTH + 5)
        self.assertEqual(
            gen._calculate_column_width([value]), len(value) + 2
        )

    def test_zero_freeze_clears_an_inherited_freeze(self) -> None:
        # A template workbook may arrive with panes already frozen; freeze_rows: 0
        # must actually unfreeze rather than leave the template's value in place.
        from openpyxl import Workbook

        gen = SheetGenerator()
        ws = Workbook().active
        ws.freeze_panes = "B3"
        gen._apply_freeze_panes(ws, 0, 0)
        self.assertIsNone(ws.freeze_panes)
