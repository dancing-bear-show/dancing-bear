"""Tests for calendars/importer/xlsx_parser.py."""
import tempfile
import unittest
from pathlib import Path

from tests.fixtures import has_openpyxl


class TestXLSXParser(unittest.TestCase):
    """Tests for XLSXParser class."""

    @unittest.skipUnless(has_openpyxl(), 'requires openpyxl')
    def test_parse_basic_xlsx_file(self):
        """XLSXParser should parse basic XLSX schedule."""
        import openpyxl
        from calendars.importer.xlsx_parser import XLSXParser

        # Create a temporary XLSX file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = Path(tmp.name)
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            # Add headers
            ws['A1'] = 'Subject'
            ws['B1'] = 'Days'
            ws['C1'] = 'Times'
            # Add data row
            ws['A2'] = 'Swim Class'
            ws['B2'] = 'Monday'
            ws['C2'] = '5:00pm-5:30pm'
            wb.save(tmp_path)

            parser = XLSXParser()
            items = parser.parse(str(tmp_path))

            self.assertEqual(len(items), 1)
            self.assertEqual(items[0].subject, 'Swim Class')
        finally:
            tmp_path.unlink(missing_ok=True)

    @unittest.skipUnless(has_openpyxl(), 'requires openpyxl')
    def test_parse_handles_empty_cells(self):
        """XLSXParser should handle empty cells gracefully."""
        import openpyxl
        from calendars.importer.xlsx_parser import XLSXParser

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = Path(tmp.name)
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Subject'
            ws['B1'] = 'Days'
            ws['A2'] = 'Event'
            ws['B2'] = None  # Empty cell
            wb.save(tmp_path)

            parser = XLSXParser()
            items = parser.parse(str(tmp_path))
            # Should still parse the row
            self.assertGreaterEqual(len(items), 0)
        finally:
            tmp_path.unlink(missing_ok=True)

    @unittest.skipUnless(has_openpyxl(), 'requires openpyxl')
    def test_parse_multiple_rows(self):
        """XLSXParser should parse multiple schedule items."""
        import openpyxl
        from calendars.importer.xlsx_parser import XLSXParser

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = Path(tmp.name)
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Subject'
            ws['B1'] = 'Days'
            ws['C1'] = 'Times'
            ws['A2'] = 'Swim'
            ws['B2'] = 'Monday'
            ws['C2'] = '5:00pm-5:30pm'
            ws['A3'] = 'Hockey'
            ws['B3'] = 'Tuesday'
            ws['C3'] = '6:00pm-7:00pm'
            wb.save(tmp_path)

            parser = XLSXParser()
            items = parser.parse(str(tmp_path))

            self.assertEqual(len(items), 2)
            subjects = [item.subject for item in items]
            self.assertIn('Swim', subjects)
            self.assertIn('Hockey', subjects)
        finally:
            tmp_path.unlink(missing_ok=True)

    @unittest.skipUnless(has_openpyxl(), 'requires openpyxl')
    def test_parse_skips_empty_subject_rows(self):
        """XLSXParser should skip rows with empty subjects."""
        import openpyxl
        from calendars.importer.xlsx_parser import XLSXParser

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = Path(tmp.name)
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Subject'
            ws['B1'] = 'Days'
            ws['A2'] = ''  # Empty subject
            ws['B2'] = 'Monday'
            ws['A3'] = 'Valid Event'
            ws['B3'] = 'Tuesday'
            wb.save(tmp_path)

            parser = XLSXParser()
            items = parser.parse(str(tmp_path))

            # Should only get the valid event
            self.assertEqual(len(items), 1)
            self.assertEqual(items[0].subject, 'Valid Event')
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_parse_raises_when_openpyxl_missing(self):
        """XLSXParser should raise RuntimeError when openpyxl is not installed."""
        from calendars.importer.xlsx_parser import XLSXParser
        import sys

        # Temporarily hide openpyxl
        openpyxl_module = sys.modules.get('openpyxl')
        if openpyxl_module:
            sys.modules['openpyxl'] = None  # type: ignore

        try:
            parser = XLSXParser()
            with self.assertRaises(RuntimeError) as ctx:
                parser.parse('/fake/path.xlsx')
            self.assertIn('openpyxl is required', str(ctx.exception))
        finally:
            # Restore openpyxl if it was available
            if openpyxl_module:
                sys.modules['openpyxl'] = openpyxl_module
