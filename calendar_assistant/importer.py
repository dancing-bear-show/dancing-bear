from __future__ import annotations

"""Schedule Importer scaffolding.

Parses schedules from simple tabular sources (CSV/XLSX) and produces
recurring/one-off event specs suitable for Outlook calendar creation.

Heavy parsers (PDF/HTML) are optional and lazily imported.
"""

from dataclasses import dataclass
from typing import List, Optional, Dict, Any
import os


@dataclass
class ScheduleItem:
    subject: str
    # One-off support (ISO datetimes)
    start_iso: Optional[str] = None    # YYYY-MM-DDTHH:MM[:SS]
    end_iso: Optional[str] = None
    # Recurring support
    recurrence: Optional[str] = None   # weekly|daily|monthly
    byday: Optional[List[str]] = None  # e.g., ["MO","WE"] for weekly
    start_time: Optional[str] = None   # HH:MM (24h)
    end_time: Optional[str] = None
    range_start: Optional[str] = None  # YYYY-MM-DD
    range_until: Optional[str] = None
    count: Optional[int] = None
    # Location / notes
    location: Optional[str] = None     # Either name, or "Name (street, city, ST POSTAL)"
    notes: Optional[str] = None


def parse_csv(path: str) -> List[ScheduleItem]:
    import csv  # stdlib
    items: List[ScheduleItem] = []
    with open(path, newline='', encoding='utf-8') as fh:
        rd = csv.DictReader(fh)
        for raw in rd:
            # Normalize keys: strip spaces and lower-case header names
            row = {str(k).strip().lower(): (raw.get(k) if raw.get(k) is not None else '') for k in raw.keys()}
            subj = (row.get('subject') or row.get('Subject') or '').strip()
            if not subj:
                continue
            # ByDay can be comma-separated codes or words
            byday_raw = (row.get('byday') or row.get('ByDay') or '').strip()
            byday = None
            if byday_raw:
                byday = [s.strip().upper() for s in byday_raw.split(',') if s.strip()]
            items.append(
                ScheduleItem(
                    subject=subj,
                    start_iso=(row.get('start') or row.get('Start') or '').strip() or None,
                    end_iso=(row.get('end') or row.get('End') or '').strip() or None,
                    recurrence=((row.get('recurrence') or row.get('repeat') or row.get('Repeat') or '') or '').strip().lower() or None,
                    byday=byday,
                    start_time=(row.get('starttime') or row.get('start_time') or row.get('StartTime') or '').strip() or None,
                    end_time=(row.get('endtime') or row.get('end_time') or row.get('EndTime') or '').strip() or None,
                    range_start=(row.get('startdate') or row.get('start_date') or row.get('StartDate') or '').strip() or None,
                    range_until=(row.get('until') or row.get('enddate') or row.get('EndDate') or '').strip() or None,
                    count=int(row['count']) if (row.get('count') or '').strip().isdigit() else None,
                    location=(row.get('location') or row.get('address') or row.get('Location') or row.get('Address') or '').strip() or None,
                    notes=(row.get('notes') or row.get('Notes') or '').strip() or None,
                )
            )
    return items


def parse_xlsx(path: str) -> List[ScheduleItem]:
    try:
        import openpyxl  # type: ignore
    except Exception as e:  # pragma: no cover - optional
        raise RuntimeError("openpyxl is required to parse .xlsx files. Try: python3 -m pip install openpyxl") from e
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # Expect a header row; map columns by name
    headers: Dict[int, str] = {}
    items: List[ScheduleItem] = []
    for i, cell in enumerate(ws[1], start=1):
        headers[i] = str(cell.value or '').strip()
    for r in ws.iter_rows(min_row=2):
        row: Dict[str, Any] = {}
        for i, cell in enumerate(r, start=1):
            row[headers.get(i, f'col{i}')] = cell.value if cell.value is not None else ''
        subj = str(row.get('Subject') or row.get('subject') or '').strip()
        if not subj:
            continue
        byday_raw = str(row.get('ByDay') or row.get('byday') or '').strip()
        byday = [s.strip().upper() for s in byday_raw.split(',') if s.strip()] if byday_raw else None
        items.append(
            ScheduleItem(
                subject=subj,
                start_iso=str(row.get('Start') or row.get('start') or '').strip() or None,
                end_iso=str(row.get('End') or row.get('end') or '').strip() or None,
                recurrence=str(row.get('Recurrence') or row.get('repeat') or row.get('Repeat') or '').strip().lower() or None,
                byday=byday,
                start_time=str(row.get('StartTime') or row.get('start_time') or '').strip() or None,
                end_time=str(row.get('EndTime') or row.get('end_time') or '').strip() or None,
                range_start=str(row.get('StartDate') or row.get('start_date') or '').strip() or None,
                range_until=str(row.get('Until') or row.get('EndDate') or row.get('until') or '').strip() or None,
                count=int(row['Count']) if str(row.get('Count') or '').strip().isdigit() else None,
                location=str(row.get('Location') or row.get('location') or row.get('Address') or row.get('address') or '').strip() or None,
                notes=str(row.get('Notes') or row.get('notes') or '').strip() or None,
            )
        )
    return items


def parse_pdf(path: str) -> List[ScheduleItem]:  # scaffold
    try:
        from pdfminer.high_level import extract_text  # type: ignore
    except Exception as e:  # pragma: no cover - optional
        raise RuntimeError("pdfminer.six is required to parse PDFs. Try: python3 -m pip install pdfminer.six") from e
    items: List[ScheduleItem] = []
    # First, attempt table extraction where possible for structured schedules
    try:
        import pdfplumber  # type: ignore
    except Exception:
        pdfplumber = None  # type: ignore
    if pdfplumber is not None:
        try:
            import re as _re
            with pdfplumber.open(str(path)) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables() or []
                    for tbl in tables:
                        if not tbl or not isinstance(tbl, list):
                            continue
                        header = [str(x or '').strip() for x in (tbl[0] or [])]
                        # Identify Day and Leisure Swim columns
                        day_idx = None
                        leisure_idx = None
                        for i, h in enumerate(header):
                            hl = h.replace('\n',' ').lower()
                            if day_idx is None and 'day' in hl:
                                day_idx = i
                            if leisure_idx is None and 'leisure' in hl:
                                leisure_idx = i
                        if day_idx is None or leisure_idx is None:
                            continue
                        # Parse rows
                        for row in tbl[1:]:
                            if not row or len(row) <= max(day_idx, leisure_idx):
                                continue
                            day_spec = str(row[day_idx] or '').replace('\n',' ').strip()
                            leisure_cell = str(row[leisure_idx] or '').replace('\n',' ').strip()
                            if not leisure_cell:
                                continue
                            # Normalize day tokens to weekday codes
                            def _norm_days(spec: str) -> List[str]:
                                s = (spec or '').lower().replace('&',' & ').replace('to',' to ')
                                days = ['mon','tue','wed','thu','fri','sat','sun']
                                code = {'mon':'MO','tue':'TU','wed':'WE','thu':'TH','fri':'FR','sat':'SA','sun':'SU'}
                                out: List[str] = []
                                m = _re.search(r'\b(mon|tue|wed|thu|fri|sat|sun)\b\s*(?:-|to)\s*\b(mon|tue|wed|thu|fri|sat|sun)\b', s)
                                if m:
                                    a, b = m.group(1), m.group(2)
                                    i1, i2 = days.index(a), days.index(b)
                                    rng = days[i1:i2+1] if i1 <= i2 else (days[i1:]+days[:i2+1])
                                    return [code[d] for d in rng]
                                for d in days:
                                    if _re.search(rf'\b{d}(?:day)?\b', s):
                                        c = code[d]
                                        if c not in out:
                                            out.append(c)
                                return out
                            # Extract one or more time ranges from the leisure cell
                            def _ranges(cell: str) -> List[tuple[str,str]]:
                                s = (cell or '').replace('*',' ')
                                outs: List[tuple[str,str]] = []
                                for m in _re.finditer(r'(\d{1,2}(?::\d{2})?\s*(?:a\.?m\.?|p\.?m\.?))\s*(?:-|to)\s*(\d{1,2}(?::\d{2})?\s*(?:a\.?m\.?|p\.?m\.?))', s, _re.I):
                                    a, b = m.group(1), m.group(2)
                                    def to24(x: str) -> str:
                                        x = x.strip().lower().replace(' ', '')
                                        mm = _re.match(r'^(\d{1,2})(?::(\d{2}))?((?:a\.?m\.?|p\.?m\.?))$', x)
                                        hh = int(mm.group(1)); mi = int(mm.group(2) or 0); ap = mm.group(3)
                                        amp = ap[0]  # 'a' or 'p'
                                        if amp == 'p' and hh < 12:
                                            hh += 12
                                        if amp == 'a' and hh == 12:
                                            hh = 0
                                        return f"{hh:02d}:{mi:02d}"
                                    outs.append((to24(a), to24(b)))
                                return outs
                            for code in _norm_days(day_spec):
                                for st, en in _ranges(leisure_cell):
                                    items.append(ScheduleItem(
                                        subject='Leisure Swim',
                                        recurrence='weekly', byday=[code], start_time=st, end_time=en,
                                        range_start=_dt.date.today().isoformat(), location='Aurora Pools', notes=f'Imported from PDF {path}',
                                    ))
        except Exception:
            # If pdfplumber fails, continue with text fallback below
            pass
        if items:
            return items
    try:
        text = extract_text(str(path))
    except Exception as e:
        # Surface a simple error
        raise RuntimeError(f"Failed to extract text from PDF: {e}")
    # Target known Aurora Aquatics PDF shape (drop-in schedules)
    if "Town of Aurora" in text and ("Swimming Drop-In Schedules" in text or "drop-in Lane and Leisure swims" in text):
        import re
        t = re.sub(r"\r", "\n", text)
        t = re.sub(r"\n{2,}", "\n", t)
        # Helper: day spec to weekday codes
        def norm_days(spec: str) -> List[str]:
            s = (spec or '').lower().replace('&', ' & ').replace('to', ' to ')
            days = ['mon','tue','wed','thu','fri','sat','sun']
            code = {'mon':'MO','tue':'TU','wed':'WE','thu':'TH','fri':'FR','sat':'SA','sun':'SU'}
            out: List[str] = []
            m = re.search(r'\b(mon|tue|wed|thu|fri|sat|sun)\b\s*(?:-|to)\s*\b(mon|tue|wed|thu|fri|sat|sun)\b', s)
            if m:
                a, b = m.group(1), m.group(2)
                i1, i2 = days.index(a), days.index(b)
                rng = days[i1:i2+1] if i1 <= i2 else (days[i1:]+days[:i2+1])
                return [code[d] for d in rng]
            for d in days:
                if re.search(rf'\b{d}(?:day)?\b', s):
                    c = code[d]
                    if c not in out:
                        out.append(c)
            return out
        # Helper: parse time blocks like "10:00 a.m. - 12:00 p.m." or "9:00pm - 10:30pm"
        def parse_ranges(txt: str) -> List[tuple[str, str]]:
            s = (txt or '').replace('\n', ' ')
            outs: List[tuple[str,str]] = []
            for m in re.finditer(r'(\d{1,2}(?::\d{2})?\s*(?:a|p)\.m\.)\s*(?:to|-)\s*(\d{1,2}(?::\d{2})?\s*(?:a|p)\.m\.)', s, re.I):
                a, b = m.group(1), m.group(2)
                def to24(x: str) -> str:
                    x = x.strip().lower().replace(' ', '')
                    mm = re.match(r'^(\d{1,2})(?::(\d{2}))?([ap])\.m\.$', x)
                    hh = int(mm.group(1)); mi = int(mm.group(2) or 0); ap = mm.group(3)
                    if ap == 'p' and hh < 12: hh += 12
                    if ap == 'a' and hh == 12: hh = 0
                    return f"{hh:02d}:{mi:02d}"
                outs.append((to24(a), to24(b)))
            return outs
        # Find Public Skating in Aurora Skating PDF as a baseline (optional), but focus on Leisure Swim in Aquatics PDF if present.
        # Look for blocks that mention "Day" and both "Lane" and "Leisure" headers.
        # Split into rough sections by headers and process rows.
        blocks = re.split(r'\n\s*Day\s*\n', t)
        for blk in blocks[1:]:
            # Heuristic: check if block contains "Leisure" header and at least one weekday
            if not re.search(r'Leisure', blk, re.I):
                continue
            if not re.search(r'\b(Mon|Tue|Wed|Thu|Fri|Sat|Sun)', blk, re.I):
                continue
            # Split into row-ish chunks, starting with a day spec
            rows = re.findall(r'((?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)[^\n]*?(?:\bto\b[^\n]*)?)\n([\s\S]*?)(?=\n(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)|\Z)', blk, re.I)
            for day_spec, rest in rows:
                # Extract Leisure Swim times from rest
                # If the block has multiple columns, "Leisure" may be after the Day => capture all times and let the user filter manually if needed.
                for st, en in parse_ranges(rest):
                    for code in norm_days(day_spec):
                        items.append(ScheduleItem(
                            subject='Leisure Swim',
                            recurrence='weekly', byday=[code], start_time=st, end_time=en,
                            range_start=_dt.date.today().isoformat(), location='Aurora Pools', notes=f'Imported from PDF {path}',
                        ))
        return items
    # Fallback
    raise NotImplementedError("Generic PDF parsing not implemented. This parser supports Aurora drop-in schedules.")


def parse_website(url: str) -> List[ScheduleItem]:  # targeted scaffold for Richmond Hill skating
    import re
    import datetime as _dt
    import requests
    try:
        from bs4 import BeautifulSoup  # type: ignore
    except Exception:
        BeautifulSoup = None  # type: ignore

    u = str(url or '')
    if 'richmondhill.ca' in u and 'Skating.aspx' in u:
        html = requests.get(u, timeout=30).text
        # Prefer BeautifulSoup when available for robust parsing
        if BeautifulSoup is not None:
            soup = BeautifulSoup(html, 'html.parser')
            out: List[ScheduleItem] = []
            # The page uses paired rows: accParent (arena name) then accChild (table)
            parents = soup.select('[data-name="accParent"]')
            for p in parents:
                arena = (p.get_text(strip=True) or '').strip()
                sib = p.find_next(attrs={'data-name': 'accChild'})
                if not sib:
                    continue
                table = sib.find('table')
                if not table:
                    continue
                # Build day headers
                days: List[str] = []
                for td in table.find_all('td'):
                    txt = (td.get_text(strip=True) or '').strip()
                    if txt.lower() in ('sunday','monday','tuesday','wednesday','thursday','friday','saturday'):
                        days.append(txt)
                # Find row with Public Skating
                row = None
                for tr in table.find_all('tr'):
                    first = tr.find('td')
                    if first and ('public skating' in (first.get_text(strip=True) or '').lower()):
                        row = tr
                        break
                if not row:
                    continue
                # Extract cells after the first (activity name)
                cells = row.find_all('td')
                if len(cells) < 2:
                    continue
                # Build mapping of day->time string by position
                # Some tables repeat headers; pick unique order of weekday headers encountered
                if len(days) < 7:
                    # Attempt to infer headers by scanning the header row above
                    header_tr = row.find_previous('tr')
                    days = []
                    if header_tr:
                        for td in header_tr.find_all('td'):
                            t = (td.get_text(strip=True) or '').strip()
                            if t.lower() in ('sunday','monday','tuesday','wednesday','thursday','friday','saturday'):
                                days.append(t)
                # If still ambiguous, assume Sun..Sat ordering for 7 columns
                if len(days) < len(cells)-1:
                    days = ['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'][:len(cells)-1]

                def norm_day(d: str) -> str:
                    m = {
                        'sunday':'SU','monday':'MO','tuesday':'TU','wednesday':'WE','thursday':'TH','friday':'FR','saturday':'SA'
                    }
                    return m.get(d.lower(), '')

                def parse_time_range(s: str) -> tuple[Optional[str], Optional[str]]:
                    # Examples: "1:45 - 3:15 p.m.", "7:15 - 8:45 p.m.", "11:15 a.m. - 12:15 p.m.", "7 - 8:30 p.m."
                    s = (s or '').strip()
                    if not s or s == '\xa0':
                        return None, None
                    s = s.replace('\u00a0',' ').replace('\xa0',' ')
                    s = s.replace('–','-').replace('—','-').replace('to','-')
                    ampm = None
                    if 'a.m' in s.lower():
                        ampm = 'am'
                    if 'p.m' in s.lower():
                        ampm = 'pm'
                    # Strip am/pm text for split
                    s_clean = re.sub(r'(?i)\b(a\.?m\.?|p\.?m\.?)\b','',s)
                    parts = [t.strip(' .') for t in s_clean.split('-') if t.strip()]
                    if len(parts)!=2:
                        return None, None
                    def to24(t: str, suffix: Optional[str]) -> Optional[str]:
                        m = re.match(r'^(\d{1,2})(?::(\d{2}))?$', t)
                        if not m:
                            return None
                        hh = int(m.group(1))
                        mm = int(m.group(2) or 0)
                        suf = suffix
                        if suf is None:
                            # Heuristic: if hour <=7 assume pm for evening; else am
                            suf = 'pm' if hh>=7 else 'am'
                        if suf.lower().startswith('p') and hh < 12:
                            hh += 12
                        if suf.lower().startswith('a') and hh == 12:
                            hh = 0
                        return f"{hh:02d}:{mm:02d}"
                    # Try to detect am/pm for each side separately
                    ampm_left = None
                    ampm_right = None
                    if re.search(r'\b(a\.m\.|am)\b', s, re.I):
                        ampm_left = 'am'
                    if re.search(r'\b(p\.m\.|pm)\b', s, re.I):
                        ampm_right = 'pm'
                    # If only one side present, propagate
                    if ampm_left and not ampm_right:
                        ampm_right = ampm_left
                    if ampm_right and not ampm_left:
                        ampm_left = ampm_right
                    start = to24(parts[0], ampm_left or ampm)
                    end = to24(parts[1], ampm_right or ampm)
                    return start, end

                today = _dt.date.today().isoformat()
                for i, td in enumerate(cells[1:], start=0):
                    day = days[i] if i < len(days) else None
                    if not day:
                        continue
                    txt = (td.get_text(" ", strip=True) or '').strip()
                    if not txt or txt == '\xa0':
                        continue
                    st, en = parse_time_range(txt)
                    if not st or not en:
                        continue
                    out.append(
                        ScheduleItem(
                            subject="Public Skating",
                            recurrence='weekly',
                            byday=[norm_day(day)],
                            start_time=st,
                            end_time=en,
                            range_start=today,
                            location=arena,
                            notes=f"Imported from {u}",
                        )
                    )
            return out
        # Fallback: simple regex parse (best-effort)
        blocks = re.split(r'data-name="accParent"', html)
        items: List[ScheduleItem] = []
        today = _dt.date.today().isoformat()
        for b in blocks[1:]:
            mname = re.search(r'>\s*([^<]+?)\s*</td>', b)
            arena = (mname.group(1).strip() if mname else 'Arena')
            # Find the Public Skating row; then capture the next 7 td cells
            mpos = re.search(r'<td[^>]*>\s*<strong>\s*Public\s*Skating\s*</strong>\s*</td>', b, re.I | re.S)
            if not mpos:
                continue
            tail = b[mpos.end():]
            cells = re.findall(r'<td[^>]*>(.*?)</td>', tail, re.I | re.S)[:7]
            days = ['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday']
            def norm_day(d: str) -> str:
                m = {'Sunday':'SU','Monday':'MO','Tuesday':'TU','Wednesday':'WE','Thursday':'TH','Friday':'FR','Saturday':'SA'}
                return m[d]
            def strip_tags(s: str) -> str:
                return re.sub(r'<[^>]+>', '', s).replace('\xa0',' ').replace('&nbsp;',' ').strip()
            def parse_range(s: str):
                s = strip_tags(s)
                if not s:
                    return None, None
                s = s.replace('–','-').replace('—','-').replace('to','-')
                has_am = re.search(r'(?i)\b(a\.?m\.?)\b', s) is not None
                has_pm = re.search(r'(?i)\b(p\.?m\.?)\b', s) is not None
                s_clean = re.sub(r'(?i)\b(a\.?m\.?|p\.?m\.?)\b','',s)
                parts = [t.strip(' .') for t in s_clean.split('-') if t.strip()]
                if len(parts)!=2:
                    return None, None
                def to24(t, suf):
                    m=re.match(r'^(\d{1,2})(?::(\d{2}))?$', t)
                    if not m:
                        return None
                    hh=int(m.group(1)); mm=int(m.group(2) or 0)
                    if suf is None:
                        suf='pm' if hh>=7 else 'am'
                    if suf.startswith('p') and hh<12: hh+=12
                    if suf.startswith('a') and hh==12: hh=0
                    return f"{hh:02d}:{mm:02d}"
                left_suf = 'am' if (has_am and has_pm) else ('am' if has_am else ('pm' if has_pm else None))
                right_suf = 'pm' if (has_am and has_pm) else ('am' if has_am else ('pm' if has_pm else None))
                start=to24(parts[0], left_suf)
                end=to24(parts[1], right_suf)
                return start, end
            for i, cell in enumerate(cells[:7]):
                s,e = parse_range(cell)
                if not s or not e:
                    continue
                items.append(ScheduleItem(
                    subject="Public Skating",
                    recurrence='weekly', byday=[norm_day(days[i])], start_time=s, end_time=e, range_start=today, location=arena,
                    notes=f"Imported from {u}",
                ))
        return items
    if 'richmondhill.ca' in u and 'Swimming.aspx' in u:
        html = requests.get(u, timeout=30).text
        items: List[ScheduleItem] = []
        # Simple regex-driven fallback parse to avoid hard deps
        blocks = re.split(r'data-name=\"accParent\"', html)
        today = _dt.date.today().isoformat()
        # Day order appears as: Activity, Sunday..Saturday
        days = ['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday']
        def norm_day(d: str) -> str:
            m = {'Sunday':'SU','Monday':'MO','Tuesday':'TU','Wednesday':'WE','Thursday':'TH','Friday':'FR','Saturday':'SA'}
            return m.get(d, '')
        def strip_tags(s: str) -> str:
            return re.sub(r'<[^>]+>', '', s).replace('\xa0',' ').strip()
        def parse_range(s: str):
            s = strip_tags(s)
            if not s:
                return None, None
            s = s.replace('–','-').replace('—','-').replace('to','-')
            has_am = re.search(r'(?i)\b(a\.?m\.?)\b', s) is not None
            has_pm = re.search(r'(?i)\b(p\.?m\.?)\b', s) is not None
            s_clean = re.sub(r'(?i)\b(a\.?m\.?|p\.?m\.?)\b','',s)
            parts = [t.strip(' .') for t in s_clean.split('-') if t.strip()]
            if len(parts)!=2:
                return None, None
            def to24(t, suf):
                m=re.match(r'^(\d{1,2})(?::(\d{2}))?$', t)
                if not m:
                    return None
                hh=int(m.group(1)); mm=int(m.group(2) or 0)
                if suf is None:
                    suf='pm' if hh>=7 else 'am'
                if suf.startswith('p') and hh<12: hh+=12
                if suf.startswith('a') and hh==12: hh=0
                return f"{hh:02d}:{mm:02d}"
            left_suf = 'am' if (has_am and has_pm) else ('am' if has_am else ('pm' if has_pm else None))
            right_suf = 'pm' if (has_am and has_pm) else ('am' if has_am else ('pm' if has_pm else None))
            return to24(parts[0], left_suf), to24(parts[1], right_suf)
        for b in blocks[1:]:
            mname = re.search(r'>\s*([^<]+?)\s*</td>', b)
            facility = (re.sub(r'&nbsp;',' ', mname.group(1)).strip() if mname else 'Pool')
            # Look for rows with activity names we care about
            for label in ('Leisure Swim','Fun N Fit','Fun N Fit & Leisure Swim'):
                mpos = re.search(rf'<td[^>]*>\s*<strong>\s*{label}\s*</strong>\s*</td>', b, re.I | re.S)
                if not mpos:
                    continue
                tail = b[mpos.end():]
                cells = re.findall(r'<td[^>]*>(.*?)</td>', tail, re.I | re.S)[:7]
                for i, cell in enumerate(cells):
                    st,en = parse_range(cell)
                    if not st or not en:
                        continue
                    subj = 'Leisure Swim' if 'Leisure' in label and 'Fun' not in label else 'Fun N Fit'
                    items.append(ScheduleItem(
                        subject=subj,
                        recurrence='weekly', byday=[norm_day(days[i])], start_time=st, end_time=en,
                        range_start=today, location=facility, notes=f"Imported from {u}",
                    ))
        return items
    if 'aurora.ca' in u and 'aquatics-and-swim-programs' in u:
        html = requests.get(u, timeout=30).text
        items: List[ScheduleItem] = []
        today = _dt.date.today().isoformat()
        # Locate tables containing a "Leisure Swim" header
        tables = re.findall(r'<table[\s\S]*?</table>', html, re.I)
        def strip_tags(s: str) -> str:
            return re.sub(r'<[^>]+>', '', s).replace('\xa0',' ').replace('&nbsp;',' ').strip()
        def norm_days(spec: str) -> List[str]:
            s = (spec or '').lower().replace('&amp;','&')
            days = ['mon','tue','wed','thu','fri','sat','sun']
            code = {'mon':'MO','tue':'TU','wed':'WE','thu':'TH','fri':'FR','sat':'SA','sun':'SU'}
            out: List[str] = []
            # Ranges like Mon to Fri
            m = re.search(r'\b(mon|tue|wed|thu|fri|sat|sun)\b\s*(?:to|-)\s*\b(mon|tue|wed|thu|fri|sat|sun)\b', s)
            if m:
                a, b = m.group(1), m.group(2)
                i1, i2 = days.index(a), days.index(b)
                if i1 <= i2:
                    return [code[d] for d in days[i1:i2+1]]
                return [code[d] for d in (days[i1:]+days[:i2+1])]
            # Conjunctions like Mon & Wed
            for d in days:
                if re.search(rf'\b{d}(?:day)?\b', s):
                    c = code[d]
                    if c not in out:
                        out.append(c)
            return out
        def parse_time_blocks(cell_html: str) -> List[tuple[str,str]]:
            txt = strip_tags(cell_html)
            if not txt:
                return []
            parts = re.split(r'\s*(?:\n|;|\|)\s*', txt)
            out = []
            for part in parts:
                part = part.replace('–','-').replace('—','-').replace('to','-')
                has_am = re.search(r'(?i)\b(a\.?m\.?)\b', part) is not None
                has_pm = re.search(r'(?i)\b(p\.?m\.?)\b', part) is not None
                sc = re.sub(r'(?i)\b(a\.?m\.?|p\.?m\.?)\b','', part)
                seg = [t.strip(' .') for t in sc.split('-') if t.strip()]
                if len(seg) != 2:
                    continue
                def to24(t: str, suf: str|None) -> str|None:
                    m = re.match(r'^(\d{1,2})(?::(\d{2}))?$', t)
                    if not m:
                        return None
                    hh = int(m.group(1)); mm = int(m.group(2) or 0)
                    if suf is None:
                        suf = 'pm' if hh >= 7 else 'am'
                    if suf.startswith('p') and hh < 12:
                        hh += 12
                    if suf.startswith('a') and hh == 12:
                        hh = 0
                    return f"{hh:02d}:{mm:02d}"
                left = 'am' if (has_am and has_pm) else ('am' if has_am else ('pm' if has_pm else None))
                right = 'pm' if (has_am and has_pm) else ('am' if has_am else ('pm' if has_pm else None))
                st = to24(seg[0], left); en = to24(seg[1], right)
                if st and en:
                    out.append((st, en))
            return out
        for tbl in tables:
            if not re.search(r'Leisure\s*Swim', tbl, re.I):
                continue
            # Try to find the header row
            header = re.search(r'<thead[\s\S]*?<tr[\s\S]*?>([\s\S]*?)</tr>[\s\S]*?</thead>', tbl, re.I)
            headers = []
            if header:
                headers = [strip_tags(h) for h in re.findall(r'<t[dh][^>]*>([\s\S]*?)</t[dh]>', header.group(1), re.I)]
            if not headers:
                # Try the first row as headers
                first_row = re.search(r'<tr[\s\S]*?>([\s\S]*?)</tr>', tbl, re.I)
                if first_row:
                    headers = [strip_tags(h) for h in re.findall(r'<t[dh][^>]*>([\s\S]*?)</t[dh]>', first_row.group(1), re.I)]
            # Find Day and Leisure indices
            day_idx = None; leisure_idx = None
            for i, h in enumerate(headers):
                hl = h.lower()
                if day_idx is None and 'day' in hl:
                    day_idx = i
                if leisure_idx is None and 'leisure' in hl:
                    leisure_idx = i
            if day_idx is None or leisure_idx is None:
                continue
            # Body rows
            body = re.search(r'<tbody[\s\S]*?>([\s\S]*?)</tbody>', tbl, re.I)
            rows = re.findall(r'<tr[\s\S]*?>([\s\S]*?)</tr>', body.group(1), re.I) if body else re.findall(r'<tr[\s\S]*?>([\s\S]*?)</tr>', tbl, re.I)
            for row in rows:
                cols = re.findall(r'<t[dh][^>]*>([\s\S]*?)</t[dh]>', row, re.I)
                if len(cols) <= max(day_idx, leisure_idx):
                    continue
                day_spec = strip_tags(cols[day_idx])
                leisure_cell = cols[leisure_idx]
                if not strip_tags(leisure_cell):
                    continue
                for code in norm_days(day_spec):
                    for st, en in parse_time_blocks(leisure_cell):
                        items.append(ScheduleItem(
                            subject='Leisure Swim',
                            recurrence='weekly', byday=[code], start_time=st, end_time=en,
                            range_start=today, location='Aurora Pools', notes=f'Imported from {u}',
                        ))
        return items
    raise NotImplementedError("Website parsing not implemented for this source. Provide CSV/XLSX or known site.")


def load_schedule(source: str, kind: Optional[str] = None) -> List[ScheduleItem]:
    """Load schedule items from a source path/URL.

    kind can be one of: csv, xlsx, pdf, website, auto
    """
    ext = (os.path.splitext(source)[1] or '').lower()
    k = (kind or '').strip().lower()
    if k in ('', 'auto'):
        if ext == '.csv':
            return parse_csv(source)
        if ext in ('.xlsx', '.xlsm'):
            return parse_xlsx(source)
        if ext == '.pdf':
            return parse_pdf(source)
        # Default to csv
        return parse_csv(source)
    if k == 'csv':
        return parse_csv(source)
    if k == 'xlsx':
        return parse_xlsx(source)
    if k == 'pdf':
        return parse_pdf(source)
    if k in ('website', 'html', 'url'):
        return parse_website(source)
    raise ValueError(f"Unknown schedule kind: {kind}")
