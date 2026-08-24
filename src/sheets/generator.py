"""Generate Excel spreadsheets from YAML workbook definitions."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.yamlio import load_config
from sheets.constants import (
    ALTERNATING_ROW_COLOR_1,
    ALTERNATING_ROW_COLOR_2,
    DEFAULT_ALTERNATING_ROWS,
    DEFAULT_BORDER_COLOR,
    DEFAULT_FONT_NAME,
    DEFAULT_FONT_SIZE,
    DEFAULT_FREEZE_COLS,
    DEFAULT_FREEZE_ROWS,
    DEFAULT_HEADER_BG_COLOR,
    DEFAULT_HEADER_BOLD,
    DEFAULT_HEADER_FONT_SIZE,
    DEFAULT_HEADER_TEXT_COLOR,
    DEFAULT_SHEET_NAME,
    DEFAULT_WORKBOOK_TITLE,
    INVALID_SHEET_NAME_CHARS,
    MAX_COLUMN_WIDTH,
    MAX_SHEET_NAME_LENGTH,
    MIN_COLUMN_WIDTH,
    YAML_ALTERNATING_ROWS,
    YAML_AUTHOR,
    YAML_BG_COLOR,
    YAML_BOLD,
    YAML_COLUMN_WIDTHS,
    YAML_DATE,
    YAML_FONT_SIZE,
    YAML_FREEZE_COLS,
    YAML_FREEZE_ROWS,
    YAML_HEADER_STYLE,
    YAML_HEADERS,
    YAML_ROWS,
    YAML_SHEET_NAME,
    YAML_SHEETS,
    YAML_TEXT_COLOR,
    YAML_TITLE,
)
from sheets.schema import HeaderStyle, SheetMetadata, SheetTab, SheetWorkbook


def validate_workbook(workbook_def: SheetWorkbook) -> list[str]:
    """Return a list of problems that would break .xlsx generation.

    Excel constrains sheet names in ways a YAML definition can violate, and
    openpyxl only reports them when the workbook is written: an invalid
    character raises ValueError mid-generate, and an over-long name is
    downgraded to a UserWarning that still yields a file some readers refuse
    to open. Checking up front means `sheets validate` fails on a definition
    that `sheets generate` cannot produce, instead of reporting OK and
    deferring the error to write time.

    Returns an empty list when the definition is generatable.
    """
    problems: list[str] = []
    seen: set[str] = set()

    for index, sheet in enumerate(workbook_def.sheets, start=1):
        label = f"sheet {index}"
        if not sheet.name:
            problems.append(f"{label}: name is empty")
            continue

        bad_chars = sorted(set(sheet.name) & INVALID_SHEET_NAME_CHARS)
        if bad_chars:
            problems.append(
                f"{label} ({sheet.name!r}): contains character(s) Excel forbids "
                f"in a sheet name: {' '.join(bad_chars)}"
            )
        if len(sheet.name) > MAX_SHEET_NAME_LENGTH:
            problems.append(
                f"{label} ({sheet.name!r}): name is {len(sheet.name)} characters; "
                f"Excel allows at most {MAX_SHEET_NAME_LENGTH}"
            )

        # Excel treats sheet names case-insensitively, so two tabs differing
        # only in case collide and the second silently overwrites the first.
        key = sheet.name.casefold()
        if key in seen:
            problems.append(f"{label} ({sheet.name!r}): duplicate sheet name")
        seen.add(key)

    return problems


def load_workbook_from_yaml(yaml_path: str) -> SheetWorkbook:
    """Load a workbook definition from a YAML file.

    Uses core.yamlio.load_config, which returns {} for a missing or empty
    file rather than raising. A missing file produces a SheetWorkbook with
    metadata.title=DEFAULT_WORKBOOK_TITLE and sheets=[], matching the behaviour
    of slides.generator.load_deck_from_yaml for a missing input file.
    """
    data = load_config(yaml_path)

    metadata = SheetMetadata(
        title=data.get(YAML_TITLE, DEFAULT_WORKBOOK_TITLE),
        author=data.get(YAML_AUTHOR),
        date=str(data[YAML_DATE]) if YAML_DATE in data and data[YAML_DATE] is not None else None,
    )

    sheets: list[SheetTab] = []
    for sheet_data in data.get(YAML_SHEETS, []):
        header_style: HeaderStyle | None = None
        if YAML_HEADER_STYLE in sheet_data:
            hs = sheet_data[YAML_HEADER_STYLE]
            header_style = HeaderStyle(
                bg_color=hs.get(YAML_BG_COLOR, f"#{DEFAULT_HEADER_BG_COLOR}"),
                text_color=hs.get(YAML_TEXT_COLOR, f"#{DEFAULT_HEADER_TEXT_COLOR}"),
                bold=hs.get(YAML_BOLD, DEFAULT_HEADER_BOLD),
                font_size=hs.get(YAML_FONT_SIZE, DEFAULT_HEADER_FONT_SIZE),
            )

        sheet = SheetTab(
            name=sheet_data.get(YAML_SHEET_NAME, DEFAULT_SHEET_NAME),
            headers=sheet_data.get(YAML_HEADERS, []),
            rows=sheet_data.get(YAML_ROWS, []),
            header_style=header_style,
            alternating_rows=sheet_data.get(YAML_ALTERNATING_ROWS, DEFAULT_ALTERNATING_ROWS),
            column_widths=sheet_data.get(YAML_COLUMN_WIDTHS),
            freeze_rows=sheet_data.get(YAML_FREEZE_ROWS, DEFAULT_FREEZE_ROWS),
            freeze_cols=sheet_data.get(YAML_FREEZE_COLS, DEFAULT_FREEZE_COLS),
        )
        sheets.append(sheet)

    return SheetWorkbook(metadata=metadata, sheets=sheets)


class SheetGenerator:
    """Generator for Excel workbooks from SheetWorkbook definitions.

    Produces styled .xlsx files with header row fills, alternating (zebra)
    row colours, freeze panes, auto-fit column widths clamped to MIN/MAX
    bounds, and thin cell borders.
    """

    def __init__(self, template_path: str | None = None) -> None:
        """Initialise the generator.

        Args:
            template_path: Optional path to a template .xlsx file to use
                           as the base workbook instead of a blank one.
        """
        self.template_path = template_path

    def _normalize_color(self, color: str) -> str:
        """Normalise a color string to 8-character ARGB hex without ``#``.

        openpyxl expects ARGB format (e.g. ``FF2D3A4F``) for PatternFill,
        Font, and Side color fields. A 6-char RGB is assumed fully opaque
        and gets the ``FF`` alpha prefix added automatically.
        """
        if color.startswith("#"):
            color = color[1:]
        color = color.upper()
        if len(color) == 6:
            color = "FF" + color
        return color

    def _create_header_fill(self, header_style: HeaderStyle | None) -> PatternFill:
        """Return a solid PatternFill for the header row."""
        bg = self._normalize_color(header_style.bg_color) if header_style else ("FF" + DEFAULT_HEADER_BG_COLOR)
        return PatternFill(start_color=bg, end_color=bg, fill_type="solid")

    def _create_header_font(self, header_style: HeaderStyle | None) -> Font:
        """Return a Font for header cells."""
        if header_style:
            text_color = self._normalize_color(header_style.text_color)
            bold = header_style.bold
            font_size = header_style.font_size
        else:
            text_color = "FF" + DEFAULT_HEADER_TEXT_COLOR
            bold = DEFAULT_HEADER_BOLD
            font_size = DEFAULT_HEADER_FONT_SIZE
        return Font(name=DEFAULT_FONT_NAME, size=font_size, bold=bold, color=text_color)

    def _create_data_font(self) -> Font:
        """Return a Font for data cells."""
        return Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)

    def _create_border(self) -> Border:
        """Return a thin Border for all four cell edges."""
        side = Side(style="thin", color=self._normalize_color(DEFAULT_BORDER_COLOR))
        return Border(left=side, right=side, top=side, bottom=side)

    def _create_row_fill(self, row_index: int, alternating: bool) -> PatternFill | None:
        """Return a PatternFill for a data row, or None when alternating is off.

        Args:
            row_index: 0-indexed row number within the data section.
            alternating: Whether zebra striping is enabled for this sheet.
        """
        if not alternating:
            return None
        color = ALTERNATING_ROW_COLOR_1 if row_index % 2 == 0 else ALTERNATING_ROW_COLOR_2
        return PatternFill(start_color="FF" + color, end_color="FF" + color, fill_type="solid")

    def _calculate_column_width(self, values: list[Any]) -> int:
        """Calculate a column width from its content, clamped to [MIN, MAX].

        Inspects every value (header + data), takes the longest string
        representation, adds a small readability buffer, then clamps to
        MIN_COLUMN_WIDTH..MAX_COLUMN_WIDTH.
        """
        max_len = MIN_COLUMN_WIDTH
        for value in values:
            if value is not None:
                max_len = max(max_len, len(str(value)))
        return min(max_len + 2, MAX_COLUMN_WIDTH)

    def _resolve_column_widths(self, sheet: SheetTab) -> list[int]:
        """Return a list of column widths for the sheet.

        Uses explicit ``column_widths`` from the sheet definition when
        provided and long enough; otherwise auto-calculates from content.
        """
        num_cols = len(sheet.headers)
        if sheet.column_widths and len(sheet.column_widths) >= num_cols:
            return list(sheet.column_widths[:num_cols])

        widths: list[int] = []
        for col_idx in range(num_cols):
            values: list[Any] = [sheet.headers[col_idx]]
            for row in sheet.rows:
                if col_idx < len(row):
                    values.append(row[col_idx])
            widths.append(self._calculate_column_width(values))
        return widths

    def _apply_freeze_panes(self, ws: Any, freeze_rows: int, freeze_cols: int) -> None:
        """Apply freeze panes to a worksheet.

        The freeze pane cell in Excel is the first unfrozen cell, so
        freezing 1 row and 0 columns anchors at A2.
        """
        if freeze_rows <= 0 and freeze_cols <= 0:
            return
        col_letter = get_column_letter(freeze_cols + 1) if freeze_cols > 0 else "A"
        ws.freeze_panes = f"{col_letter}{freeze_rows + 1}"

    def _write_header_row(
        self,
        ws: Any,
        headers: list[str],
        header_fill: PatternFill,
        header_font: Font,
        border: Border,
        alignment: Alignment,
    ) -> None:
        """Write the header row to a worksheet."""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = alignment

    def _write_data_rows(
        self,
        ws: Any,
        sheet: SheetTab,
        data_font: Font,
        border: Border,
        alignment: Alignment,
    ) -> None:
        """Write all data rows to a worksheet."""
        for row_idx, row_data in enumerate(sheet.rows):
            excel_row = row_idx + 2  # 1-indexed; row 1 is the header
            row_fill = self._create_row_fill(row_idx, sheet.alternating_rows)
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.font = data_font
                cell.border = border
                cell.alignment = alignment
                if row_fill is not None:
                    cell.fill = row_fill

    def _build_worksheet(self, ws: Any, sheet: SheetTab) -> None:
        """Populate a single worksheet from a SheetTab definition."""
        ws.title = sheet.name

        header_fill = self._create_header_fill(sheet.header_style)
        header_font = self._create_header_font(sheet.header_style)
        data_font = self._create_data_font()
        border = self._create_border()
        alignment = Alignment(vertical="center", wrap_text=False)

        col_widths = self._resolve_column_widths(sheet)
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        self._write_header_row(ws, sheet.headers, header_fill, header_font, border, alignment)
        self._write_data_rows(ws, sheet, data_font, border, alignment)
        self._apply_freeze_panes(ws, sheet.freeze_rows, sheet.freeze_cols)

    def generate(self, workbook_def: SheetWorkbook, output_path: str) -> str:
        """Generate an Excel file from a SheetWorkbook definition.

        Creates parent directories if they do not exist.

        Args:
            workbook_def: Parsed workbook definition.
            output_path: Destination path for the .xlsx file.

        Returns:
            The resolved output path as a string.
        """
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        if self.template_path:
            from openpyxl import load_workbook as _load_workbook
            wb = _load_workbook(self.template_path)
        else:
            wb = Workbook()

        wb.properties.title = workbook_def.metadata.title
        if workbook_def.metadata.author:
            wb.properties.creator = workbook_def.metadata.author

        for idx, sheet in enumerate(workbook_def.sheets):
            if idx == 0 and not self.template_path:
                ws = wb.active
            else:
                ws = wb.create_sheet()
            self._build_worksheet(ws, sheet)

        wb.save(output_path)
        return output_path


def generate_xlsx(workbook_def: SheetWorkbook, output_path: str) -> str:
    """Generate an Excel file from a SheetWorkbook definition.

    Convenience wrapper around ``SheetGenerator().generate()``.
    """
    return SheetGenerator().generate(workbook_def, output_path)


def generate_from_yaml(yaml_path: str, output_path: str) -> str:
    """Load a YAML workbook definition and generate an .xlsx file.

    Convenience wrapper combining ``load_workbook_from_yaml`` and
    ``SheetGenerator().generate()``.
    """
    return SheetGenerator().generate(load_workbook_from_yaml(yaml_path), output_path)
