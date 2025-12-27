"""Schedule Importer scaffolding.

Parses schedules from simple tabular sources (CSV/XLSX) and produces
recurring/one-off event specs suitable for Outlook calendar creation.

Heavy parsers (PDF/HTML) are optional and lazily imported.
"""
from __future__ import annotations

import os
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

from core.constants import DEFAULT_REQUEST_TIMEOUT
from core.text_utils import html_to_text, normalize_unicode

from .scan_common import DAY_MAP

# Regex pattern constants for HTML parsing
RE_STRIP_TAGS = r'<[^>]+>'
RE_AMPM = r'(?i)\b(a\.?m\.?|p\.?m\.?)\b'
RE_AM_ONLY = r'(?i)\b(a\.?m\.?)\b'
RE_PM_ONLY = r'(?i)\b(p\.?m\.?)\b'
RE_TIME = r'^(\d{1,2})(?::(\d{2}))?'
RE_TABLE_CELL = r'<t[dh][^>]*>([\s\S]*?)</t[dh]>'
RE_TABLE_ROW = r'<tr[\s\S]*?>([\s\S]*?)</tr>'

# Day name sequence for iteration
DAY_NAMES = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']


def normalize_day(day_name: str) -> str:
    """Convert day name to two-letter code (e.g., 'Monday' -> 'MO')."""
    return DAY_MAP.get(day_name.lower().strip(), '')


def normalize_days(spec: str) -> List[str]:
    """Parse day specification to list of two-letter codes.

    Handles ranges like 'Mon to Fri' and lists like 'Mon & Wed'.
    Also handles full day names like 'Monday', 'Saturday'.
    """
    import re
    s = (spec or '').lower().replace('&', ' & ').replace('to', ' to ').replace('&amp;', '&')
    out: List[str] = []

    # Check for ranges like "Mon to Fri" or "Mon-Fri"
    m = re.search(r'\b(mon|tue|wed|thu|fri|sat|sun)\w*\b\s*(?:-|to)\s*\b(mon|tue|wed|thu|fri|sat|sun)\w*\b', s)
    if m:
        a, b = m.group(1), m.group(2)
        i1, i2 = DAY_NAMES.index(a), DAY_NAMES.index(b)
        rng = DAY_NAMES[i1:i2+1] if i1 <= i2 else (DAY_NAMES[i1:] + DAY_NAMES[:i2+1])
        return [DAY_MAP[d] for d in rng]

    # Check for individual days (supports both abbreviated and full names)
    for d in DAY_NAMES:
        if re.search(rf'\b{d}\w*\b', s):
            c = DAY_MAP[d]
            if c not in out:
                out.append(c)
    return out


def to_24h(time_str: str, am_pm: Optional[str] = None) -> Optional[str]:
    """Convert time string to 24-hour format (e.g., '1:45 p.m.' -> '13:45').

    If am_pm is None and not detectable, uses heuristic: hour >= 7 assumes PM.
    """
    import re
    t = (time_str or '').strip().lower().replace(' ', '')

    # Try to extract am/pm from the string itself
    detected_ampm = am_pm
    if detected_ampm is None:
        if re.search(r'p\.?m\.?', t):
            detected_ampm = 'pm'
        elif re.search(r'a\.?m\.?', t):
            detected_ampm = 'am'

    # Remove am/pm markers
    t_clean = re.sub(r'[ap]\.?m\.?', '', t).strip(' .')

    m = re.match(r'^(\d{1,2})(?::(\d{2}))?$', t_clean)
    if not m:
        return None

    hh = int(m.group(1))
    mm = int(m.group(2) or 0)

    # Apply am/pm conversion
    suf = detected_ampm
    if suf is None:
        # Heuristic: if hour >= 7 and <= 11, assume PM for evening schedules
        suf = 'pm' if 7 <= hh <= 11 else 'am'

    if suf.startswith('p') and hh < 12:
        hh += 12
    if suf.startswith('a') and hh == 12:
        hh = 0

    return f"{hh:02d}:{mm:02d}"


def parse_time_range(s: str) -> tuple[Optional[str], Optional[str]]:
    """Parse time range string to (start_24h, end_24h) tuple.

    Examples: '1:45 - 3:15 p.m.', '11:15 a.m. - 12:15 p.m.', '7 - 8:30 p.m.'
    """
    import re
    s = (s or '').strip()
    if not s or s == '\xa0':
        return None, None

    # Normalize unicode and separators
    s = normalize_unicode(s).replace(' to ', '-')

    # Detect am/pm for each side
    has_am = re.search(RE_AM_ONLY, s) is not None
    has_pm = re.search(RE_PM_ONLY, s) is not None

    # Strip am/pm text for splitting
    s_clean = re.sub(RE_AMPM, '', s)
    parts = [t.strip(' .') for t in s_clean.split('-') if t.strip()]
    if len(parts) != 2:
        return None, None

    # Determine am/pm for left and right sides
    if has_am and has_pm:
        left_suf, right_suf = 'am', 'pm'
    elif has_am:
        left_suf = right_suf = 'am'
    elif has_pm:
        left_suf = right_suf = 'pm'
    else:
        left_suf = right_suf = None

    start = to_24h(parts[0], left_suf)
    end = to_24h(parts[1], right_suf)
    return start, end


def extract_time_ranges(text: str) -> List[tuple[str, str]]:
    """Extract all time ranges from text.

    Finds patterns like '10:00 a.m. - 12:00 p.m.' or '9:00pm - 10:30pm'.
    Returns list of (start_24h, end_24h) tuples.
    """
    import re
    text = (text or '').replace('*', ' ').replace('\n', ' ')
    results: List[tuple[str, str]] = []

    # Match time range patterns with am/pm
    pattern = r'(\d{1,2}(?::\d{2})?\s*(?:a\.?m\.?|p\.?m\.?))\s*(?:-|to)\s*(\d{1,2}(?::\d{2})?\s*(?:a\.?m\.?|p\.?m\.?))'
    for m in re.finditer(pattern, text, re.I):
        start = to_24h(m.group(1))
        end = to_24h(m.group(2))
        if start and end:
            results.append((start, end))

    return results


@dataclass
class ScheduleItem:
    subject: str
    # One-off support (ISO datetimes)
    start_iso: Optional[str] = None    # YYYY-MM-DDTHH:MM[:SS]
    end_iso: Optional[str] = None
    # Recurring support
    recurrence: Optional[str] = None   # weekly|daily|monthly
    byday: Optional[List[str]] = None  # e.g., ["MO","WE"] for weekly
    start_time: Optional[str] = None   # HH:MM (24h)
    end_time: Optional[str] = None
    range_start: Optional[str] = None  # YYYY-MM-DD
    range_until: Optional[str] = None
    count: Optional[int] = None
    # Location / notes
    location: Optional[str] = None     # Either name, or "Name (street, city, ST POSTAL)"
    notes: Optional[str] = None


def parse_csv(path: str) -> List[ScheduleItem]:
    import csv  # stdlib
    items: List[ScheduleItem] = []
    with open(path, newline='', encoding='utf-8') as fh:
        rd = csv.DictReader(fh)
        for raw in rd:
            # Normalize keys: strip spaces and lower-case header names
            row = {str(k).strip().lower(): (raw.get(k) if raw.get(k) is not None else '') for k in raw.keys()}
            subj = (row.get('subject') or row.get('Subject') or '').strip()
            if not subj:
                continue
            # ByDay can be comma-separated codes or words
            byday_raw = (row.get('byday') or row.get('ByDay') or '').strip()
            byday = None
            if byday_raw:
                byday = [s.strip().upper() for s in byday_raw.split(',') if s.strip()]
            items.append(
                ScheduleItem(
                    subject=subj,
                    start_iso=(row.get('start') or row.get('Start') or '').strip() or None,
                    end_iso=(row.get('end') or row.get('End') or '').strip() or None,
                    recurrence=((row.get('recurrence') or row.get('repeat') or row.get('Repeat') or '') or '').strip().lower() or None,
                    byday=byday,
                    start_time=(row.get('starttime') or row.get('start_time') or row.get('StartTime') or '').strip() or None,
                    end_time=(row.get('endtime') or row.get('end_time') or row.get('EndTime') or '').strip() or None,
                    range_start=(row.get('startdate') or row.get('start_date') or row.get('StartDate') or '').strip() or None,
                    range_until=(row.get('until') or row.get('enddate') or row.get('EndDate') or '').strip() or None,
                    count=int(row['count']) if (row.get('count') or '').strip().isdigit() else None,
                    location=(row.get('location') or row.get('address') or row.get('Location') or row.get('Address') or '').strip() or None,
                    notes=(row.get('notes') or row.get('Notes') or '').strip() or None,
                )
            )
    return items


def parse_xlsx(path: str) -> List[ScheduleItem]:
    try:
        import openpyxl  # type: ignore
    except Exception as e:  # pragma: no cover - optional
        raise RuntimeError("openpyxl is required to parse .xlsx files. Try: python3 -m pip install openpyxl") from e
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # Expect a header row; map columns by name
    headers: Dict[int, str] = {}
    items: List[ScheduleItem] = []
    for i, cell in enumerate(ws[1], start=1):
        headers[i] = str(cell.value or '').strip()
    for r in ws.iter_rows(min_row=2):
        row: Dict[str, Any] = {}
        for i, cell in enumerate(r, start=1):
            row[headers.get(i, f'col{i}')] = cell.value if cell.value is not None else ''
        subj = str(row.get('Subject') or row.get('subject') or '').strip()
        if not subj:
            continue
        byday_raw = str(row.get('ByDay') or row.get('byday') or '').strip()
        byday = [s.strip().upper() for s in byday_raw.split(',') if s.strip()] if byday_raw else None
        items.append(
            ScheduleItem(
                subject=subj,
                start_iso=str(row.get('Start') or row.get('start') or '').strip() or None,
                end_iso=str(row.get('End') or row.get('end') or '').strip() or None,
                recurrence=str(row.get('Recurrence') or row.get('repeat') or row.get('Repeat') or '').strip().lower() or None,
                byday=byday,
                start_time=str(row.get('StartTime') or row.get('start_time') or '').strip() or None,
                end_time=str(row.get('EndTime') or row.get('end_time') or '').strip() or None,
                range_start=str(row.get('StartDate') or row.get('start_date') or '').strip() or None,
                range_until=str(row.get('Until') or row.get('EndDate') or row.get('until') or '').strip() or None,
                count=int(row['Count']) if str(row.get('Count') or '').strip().isdigit() else None,
                location=str(row.get('Location') or row.get('location') or row.get('Address') or row.get('address') or '').strip() or None,
                notes=str(row.get('Notes') or row.get('notes') or '').strip() or None,
            )
        )
    return items


def parse_pdf(path: str) -> List[ScheduleItem]:  # scaffold
    import datetime as _dt
    try:
        from pdfminer.high_level import extract_text  # type: ignore
    except Exception as e:  # pragma: no cover - optional
        raise RuntimeError("pdfminer.six is required to parse PDFs. Try: python3 -m pip install pdfminer.six") from e
    items: List[ScheduleItem] = []
    # First, attempt table extraction where possible for structured schedules
    try:
        import pdfplumber  # type: ignore
    except Exception:
        pdfplumber = None  # type: ignore
    if pdfplumber is not None:
        try:
            with pdfplumber.open(str(path)) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables() or []
                    for tbl in tables:
                        if not tbl or not isinstance(tbl, list):
                            continue
                        header = [str(x or '').strip() for x in (tbl[0] or [])]
                        # Identify Day and Leisure Swim columns
                        day_idx = None
                        leisure_idx = None
                        for i, h in enumerate(header):
                            hl = h.replace('\n',' ').lower()
                            if day_idx is None and 'day' in hl:
                                day_idx = i
                            if leisure_idx is None and 'leisure' in hl:
                                leisure_idx = i
                        if day_idx is None or leisure_idx is None:
                            continue
                        # Parse rows
                        for row in tbl[1:]:
                            if not row or len(row) <= max(day_idx, leisure_idx):
                                continue
                            day_spec = str(row[day_idx] or '').replace('\n', ' ').strip()
                            leisure_cell = str(row[leisure_idx] or '').replace('\n', ' ').strip()
                            if not leisure_cell:
                                continue
                            for code in normalize_days(day_spec):
                                for st, en in extract_time_ranges(leisure_cell):
                                    items.append(ScheduleItem(
                                        subject='Leisure Swim',
                                        recurrence='weekly', byday=[code], start_time=st, end_time=en,
                                        range_start=_dt.date.today().isoformat(), location='Aurora Pools', notes=f'Imported from PDF {path}',
                                    ))
        except Exception:  # noqa: S110 - pdfplumber failure falls through to text extraction
            pass
        if items:
            return items
    try:
        text = extract_text(str(path))
    except Exception as e:
        # Surface a simple error
        raise RuntimeError(f"Failed to extract text from PDF: {e}")
    # Target known Aurora Aquatics PDF shape (drop-in schedules)
    if "Town of Aurora" in text and ("Swimming Drop-In Schedules" in text or "drop-in Lane and Leisure swims" in text):
        import re
        t = re.sub(r"\r", "\n", text)
        t = re.sub(r"\n{2,}", "\n", t)
        # Split into rough sections by headers and process rows
        blocks = re.split(r'\n\s*Day\s*\n', t)
        for blk in blocks[1:]:
            # Heuristic: check if block contains "Leisure" header and at least one weekday
            if not re.search(r'Leisure', blk, re.I):
                continue
            if not re.search(r'\b(Mon|Tue|Wed|Thu|Fri|Sat|Sun)', blk, re.I):
                continue
            # Split into row-ish chunks, starting with a day spec
            rows = re.findall(r'((?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)[^\n]*?(?:\bto\b[^\n]*)?)\n([\s\S]*?)(?=\n(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)|\Z)', blk, re.I)
            for day_spec, rest in rows:
                # Extract Leisure Swim times from rest
                for st, en in extract_time_ranges(rest):
                    for code in normalize_days(day_spec):
                        items.append(ScheduleItem(
                            subject='Leisure Swim',
                            recurrence='weekly', byday=[code], start_time=st, end_time=en,
                            range_start=_dt.date.today().isoformat(), location='Aurora Pools', notes=f'Imported from PDF {path}',
                        ))
        return items
    # Fallback
    raise NotImplementedError("Generic PDF parsing not implemented. This parser supports Aurora drop-in schedules.")


def parse_website(url: str) -> List[ScheduleItem]:  # targeted scaffold for Richmond Hill skating
    import re
    import datetime as _dt
    import requests
    try:
        from bs4 import BeautifulSoup  # type: ignore
    except Exception:
        BeautifulSoup = None  # type: ignore

    u = str(url or '')
    if 'richmondhill.ca' in u and 'Skating.aspx' in u:
        html = requests.get(u, timeout=DEFAULT_REQUEST_TIMEOUT).text
        # Prefer BeautifulSoup when available for robust parsing
        if BeautifulSoup is not None:
            soup = BeautifulSoup(html, 'html.parser')
            out: List[ScheduleItem] = []
            # The page uses paired rows: accParent (arena name) then accChild (table)
            parents = soup.select('[data-name="accParent"]')
            for p in parents:
                arena = (p.get_text(strip=True) or '').strip()
                sib = p.find_next(attrs={'data-name': 'accChild'})
                if not sib:
                    continue
                table = sib.find('table')
                if not table:
                    continue
                # Build day headers
                days: List[str] = []
                for td in table.find_all('td'):
                    txt = (td.get_text(strip=True) or '').strip()
                    if txt.lower() in ('sunday','monday','tuesday','wednesday','thursday','friday','saturday'):
                        days.append(txt)
                # Find row with Public Skating
                row = None
                for tr in table.find_all('tr'):
                    first = tr.find('td')
                    if first and ('public skating' in (first.get_text(strip=True) or '').lower()):
                        row = tr
                        break
                if not row:
                    continue
                # Extract cells after the first (activity name)
                cells = row.find_all('td')
                if len(cells) < 2:
                    continue
                # Build mapping of day->time string by position
                # Some tables repeat headers; pick unique order of weekday headers encountered
                if len(days) < 7:
                    # Attempt to infer headers by scanning the header row above
                    header_tr = row.find_previous('tr')
                    days = []
                    if header_tr:
                        for td in header_tr.find_all('td'):
                            t = (td.get_text(strip=True) or '').strip()
                            if t.lower() in ('sunday','monday','tuesday','wednesday','thursday','friday','saturday'):
                                days.append(t)
                # If still ambiguous, assume Sun..Sat ordering for 7 columns
                if len(days) < len(cells)-1:
                    days = ['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'][:len(cells)-1]

                today = _dt.date.today().isoformat()
                for i, td in enumerate(cells[1:], start=0):
                    day = days[i] if i < len(days) else None
                    if not day:
                        continue
                    txt = (td.get_text(" ", strip=True) or '').strip()
                    if not txt or txt == '\xa0':
                        continue
                    st, en = parse_time_range(txt)
                    if not st or not en:
                        continue
                    out.append(
                        ScheduleItem(
                            subject="Public Skating",
                            recurrence='weekly',
                            byday=[normalize_day(day)],
                            start_time=st,
                            end_time=en,
                            range_start=today,
                            location=arena,
                            notes=f"Imported from {u}",
                        )
                    )
            return out
        # Fallback: simple regex parse (best-effort)
        blocks = re.split(r'data-name="accParent"', html)
        items: List[ScheduleItem] = []
        today = _dt.date.today().isoformat()
        day_list = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        for b in blocks[1:]:
            mname = re.search(r'>\s*([^<]+?)\s*</td>', b)
            arena = (mname.group(1).strip() if mname else 'Arena')
            # Find the Public Skating row; then capture the next 7 td cells
            mpos = re.search(r'<td[^>]*>\s*<strong>\s*Public\s*Skating\s*</strong>\s*</td>', b, re.I | re.S)
            if not mpos:
                continue
            tail = b[mpos.end():]
            cells = re.findall(r'<td[^>]*>(.*?)</td>', tail, re.I | re.S)[:7]
            for i, cell in enumerate(cells[:7]):
                cell_text = html_to_text(cell)
                st, en = parse_time_range(cell_text)
                if not st or not en:
                    continue
                items.append(ScheduleItem(
                    subject="Public Skating",
                    recurrence='weekly', byday=[normalize_day(day_list[i])], start_time=st, end_time=en, range_start=today, location=arena,
                    notes=f"Imported from {u}",
                ))
        return items
    if 'richmondhill.ca' in u and 'Swimming.aspx' in u:
        html = requests.get(u, timeout=DEFAULT_REQUEST_TIMEOUT).text
        items: List[ScheduleItem] = []
        # Simple regex-driven fallback parse to avoid hard deps
        blocks = re.split(r'data-name=\"accParent\"', html)
        today = _dt.date.today().isoformat()
        day_list = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        for b in blocks[1:]:
            mname = re.search(r'>\s*([^<]+?)\s*</td>', b)
            facility = (re.sub(r'&nbsp;', ' ', mname.group(1)).strip() if mname else 'Pool')
            # Look for rows with activity names we care about
            for label in ('Leisure Swim', 'Fun N Fit', 'Fun N Fit & Leisure Swim'):
                mpos = re.search(rf'<td[^>]*>\s*<strong>\s*{label}\s*</strong>\s*</td>', b, re.I | re.S)
                if not mpos:
                    continue
                tail = b[mpos.end():]
                cells = re.findall(r'<td[^>]*>(.*?)</td>', tail, re.I | re.S)[:7]
                for i, cell in enumerate(cells):
                    cell_text = html_to_text(cell)
                    st, en = parse_time_range(cell_text)
                    if not st or not en:
                        continue
                    subj = 'Leisure Swim' if 'Leisure' in label and 'Fun' not in label else 'Fun N Fit'
                    items.append(ScheduleItem(
                        subject=subj,
                        recurrence='weekly', byday=[normalize_day(day_list[i])], start_time=st, end_time=en,
                        range_start=today, location=facility, notes=f"Imported from {u}",
                    ))
        return items
    if 'aurora.ca' in u and 'aquatics-and-swim-programs' in u:
        html = requests.get(u, timeout=DEFAULT_REQUEST_TIMEOUT).text
        items: List[ScheduleItem] = []
        today = _dt.date.today().isoformat()
        # Locate tables containing a "Leisure Swim" header
        tables = re.findall(r'<table[\s\S]*?</table>', html, re.I)
        for tbl in tables:
            if not re.search(r'Leisure\s*Swim', tbl, re.I):
                continue
            # Try to find the header row
            header = re.search(r'<thead[\s\S]*?<tr[\s\S]*?>([\s\S]*?)</tr>[\s\S]*?</thead>', tbl, re.I)
            headers = []
            if header:
                headers = [html_to_text(h) for h in re.findall(RE_TABLE_CELL, header.group(1), re.I)]
            if not headers:
                # Try the first row as headers
                first_row = re.search(RE_TABLE_ROW, tbl, re.I)
                if first_row:
                    headers = [html_to_text(h) for h in re.findall(RE_TABLE_CELL, first_row.group(1), re.I)]
            # Find Day and Leisure indices
            day_idx = None
            leisure_idx = None
            for i, h in enumerate(headers):
                hl = h.lower()
                if day_idx is None and 'day' in hl:
                    day_idx = i
                if leisure_idx is None and 'leisure' in hl:
                    leisure_idx = i
            if day_idx is None or leisure_idx is None:
                continue
            # Body rows
            body = re.search(r'<tbody[\s\S]*?>([\s\S]*?)</tbody>', tbl, re.I)
            rows = re.findall(RE_TABLE_ROW, body.group(1), re.I) if body else re.findall(RE_TABLE_ROW, tbl, re.I)
            for row in rows:
                cols = re.findall(RE_TABLE_CELL, row, re.I)
                if len(cols) <= max(day_idx, leisure_idx):
                    continue
                day_spec = html_to_text(cols[day_idx])
                leisure_cell = cols[leisure_idx]
                if not html_to_text(leisure_cell):
                    continue
                for code in normalize_days(day_spec):
                    for st, en in extract_time_ranges(leisure_cell):
                        items.append(ScheduleItem(
                            subject='Leisure Swim',
                            recurrence='weekly', byday=[code], start_time=st, end_time=en,
                            range_start=today, location='Aurora Pools', notes=f'Imported from {u}',
                        ))
        return items
    raise NotImplementedError("Website parsing not implemented for this source. Provide CSV/XLSX or known site.")


def load_schedule(source: str, kind: Optional[str] = None) -> List[ScheduleItem]:
    """Load schedule items from a source path/URL.

    kind can be one of: csv, xlsx, pdf, website, auto
    """
    ext = (os.path.splitext(source)[1] or '').lower()
    k = (kind or '').strip().lower()
    if k in ('', 'auto'):
        if ext == '.csv':
            return parse_csv(source)
        if ext in ('.xlsx', '.xlsm'):
            return parse_xlsx(source)
        if ext == '.pdf':
            return parse_pdf(source)
        # Default to csv
        return parse_csv(source)
    if k == 'csv':
        return parse_csv(source)
    if k == 'xlsx':
        return parse_xlsx(source)
    if k == 'pdf':
        return parse_pdf(source)
    if k in ('website', 'html', 'url'):
        return parse_website(source)
    raise ValueError(f"Unknown schedule kind: {kind}")
